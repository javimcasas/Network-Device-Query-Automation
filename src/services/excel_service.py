"""Servicio para gestionar operaciones con Excel."""
import pandas as pd
from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..config.constants import EXCEL_PATH, EXCEL_COLUMNS, DATA_DIR
from ..models.device import Device


class ExcelService:
    """Gestiona todas las operaciones relacionadas con Excel."""
    
    def __init__(self, excel_path: Path = EXCEL_PATH):
        """Inicializa el servicio con la ruta del archivo Excel."""
        self.excel_path = excel_path
        self.data_dir = DATA_DIR
        
        # Asegurar que existe el directorio data
        self.data_dir.mkdir(exist_ok=True)
    
    def exists(self) -> bool:
        """Verifica si el archivo Excel existe."""
        return self.excel_path.exists()
    
    def create(self) -> None:
        """Crea un nuevo archivo Excel con formato bonito."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Devices"
        
        # Estilo del encabezado
        header_fill = PatternFill(start_color="4472C4", 
                                   end_color="4472C4", 
                                   fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Borde
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Escribir encabezados
        for col_num, column_name in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Ajustar anchos de columna
        column_widths = {
            'A': 25,  # Name
            'B': 15,  # User
            'C': 20,  # Password
            'D': 30   # Parameter
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Congelar primera fila
        ws.freeze_panes = "A2"
        
        # Guardar
        wb.save(self.excel_path)
        print(f"✓ Excel creado exitosamente en: {self.excel_path}")
    
    def delete(self) -> None:
        """Elimina el archivo Excel si existe."""
        if self.exists():
            self.excel_path.unlink()
            print(f"✓ Excel eliminado: {self.excel_path}")
        else:
            print("⚠ El archivo Excel no existe")
    
    def read_devices(self) -> List[Device]:
        """Lee los dispositivos del Excel y los retorna como objetos Device."""
        if not self.exists():
            raise FileNotFoundError(
                f"El archivo Excel no existe: {self.excel_path}"
            )
        
        # Leer Excel
        df = pd.read_excel(self.excel_path, sheet_name='Devices')
        
        # Convertir a lista de objetos Device
        devices = []
        for _, row in df.iterrows():
            # Saltar filas vacías
            if pd.isna(row['Name']) or row['Name'] == '':
                continue
            
            device = Device.from_dict(row.to_dict())
            devices.append(device)
        
        return devices
    
    def open_file(self) -> None:
        """Abre el archivo Excel con la aplicación predeterminada."""
        if not self.exists():
            raise FileNotFoundError(
                f"El archivo Excel no existe: {self.excel_path}"
            )
        
        import os
        import platform
        
        system = platform.system()
        
        try:
            if system == 'Windows':
                os.startfile(self.excel_path)
            elif system == 'Darwin':  # macOS
                os.system(f'open "{self.excel_path}"')
            else:  # Linux
                os.system(f'xdg-open "{self.excel_path}"')
            
            print(f"✓ Abriendo Excel: {self.excel_path}")
        except Exception as e:
            print(f"✗ Error al abrir Excel: {e}")
